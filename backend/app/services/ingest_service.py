import os
import shutil
import tempfile
from fastapi import UploadFile, HTTPException
from langchain_community.document_loaders import TextLoader, PyPDFLoader, Docx2txtLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_qdrant import QdrantVectorStore
from qdrant_client import QdrantClient
from qdrant_client.http import models
from app.core.config import settings
import logging
import re

logger = logging.getLogger(__name__)

class IngestService:
    def __init__(self):
        self.embeddings = HuggingFaceEmbeddings(model_name=settings.EMBEDDING_MODEL)
        self.client = QdrantClient(url=settings.QDRANT_URL)
        self.vector_store = QdrantVectorStore(
            client=self.client,
            collection_name=settings.COLLECTION_NAME,
            embedding=self.embeddings,
        )
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1500,
            chunk_overlap=300,
            add_start_index=True,
        )

    async def extract_text_from_file(self, file: UploadFile) -> str:
        """
        Extract text from a file without ingesting it.
        """
        try:
            suffix = os.path.splitext(file.filename)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                shutil.copyfileobj(file.file, tmp)
                tmp_path = tmp.name

            try:
                documents = []
                if suffix.lower() == ".pdf":
                    loader = PyPDFLoader(tmp_path)
                    documents = loader.load()
                elif suffix.lower() == ".docx":
                    loader = Docx2txtLoader(tmp_path)
                    documents = loader.load()
                elif suffix.lower() == ".txt":
                    loader = TextLoader(tmp_path, encoding="utf-8")
                    documents = loader.load()
                elif suffix.lower() == ".xlsx":
                    return self._extract_xlsx_text(tmp_path)
                else:
                   raise ValueError(f"Unsupported file type: {suffix}")

                if not documents:
                    return ""

                return "\n\n".join([self._clean_text(doc.page_content) for doc in documents])

            finally:
                if os.path.exists(tmp_path):
                    # Reset file cursor for subsequent operations if needed, 
                    # though here we consumed the upload file stream.
                    # Fastapi UploadFile might need seek(0) if reused.
                    pass
                    os.remove(tmp_path)

        except Exception as e:
            logger.error(f"Error extracting text from {file.filename}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to extract text: {str(e)}")

    def _clean_text(self, text: str) -> str:
        """
        Clean extracted text to fix common PDF extraction artifacts.
        e.g. "wordWord" -> "word Word", "end.Start" -> "end. Start"
        """
        if not text:
            return ""
            
        # 1. Normalization: paragraphs
        text = re.sub(r'\n\s*\n', '<PARAGRAPH>', text)
        
        # 2. Unwrap single newlines
        text = re.sub(r'\n', ' ', text)
        
        # 3. Restore paragraphs
        text = text.replace('<PARAGRAPH>', '\n\n')
            
        # 4. Fix camelCase-like merges
        text = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', text)
        
        # 5. Fix UpperTitle merges (e.g. SYSTEMSIf -> SYSTEMS If)
        text = re.sub(r'(?<=[A-Z])(?=[A-Z][a-z])', ' ', text)
        
        # 6. Fix period followed by Uppercase (e.g. ac.keHe -> ac.ke He)
        text = re.sub(r'(?<=[a-z]\.)(?=[A-Z])', ' ', text)
        
        # 7. Collapse multiple spaces
        text = re.sub(r'[ \t]+', ' ', text)
        
        return text.strip()

    async def process_file(self, file: UploadFile, user_id: str):
        """
        Process an uploaded file: save to temp, load, split, and ingest.
        """
        try:
            # Create a temporary file to save the upload
            suffix = os.path.splitext(file.filename)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                shutil.copyfileobj(file.file, tmp)
                tmp_path = tmp.name

            try:
                documents = []
                # Determine loader based on file extension
                if suffix.lower() == ".pdf":
                    loader = PyPDFLoader(tmp_path)
                    documents = loader.load()
                elif suffix.lower() == ".docx":
                    loader = Docx2txtLoader(tmp_path)
                    documents = loader.load()
                elif suffix.lower() == ".txt":
                    loader = TextLoader(tmp_path, encoding="utf-8")
                    documents = loader.load()
                elif suffix.lower() == ".xlsx":
                    from langchain_core.documents import Document
                    text = self._extract_xlsx_text(tmp_path)
                    documents = [Document(page_content=text, metadata={"source": file.filename})]
                else:
                    # For images, we would use a Vision model here.
                    # For now, we skip unsupported types or implement image handling later.
                    logger.warning(f"Unsupported file type: {suffix}")
                    return {"success": False, "message": f"Unsupported file type: {suffix}"}

                if not documents:
                    return {"success": False, "message": "No content found in file."}

                # Add metadata and Clean Text
                for doc in documents:
                    doc.page_content = self._clean_text(doc.page_content)
                    doc.metadata["source"] = file.filename
                    doc.metadata["user_id"] = user_id
                    doc.metadata["type"] = "upload"

                # Split text
                texts = self.text_splitter.split_documents(documents)
                
                if not texts:
                     return {"success": False, "message": "Could not split documents."}

                # Ingest into Qdrant
                logger.info(f"Adding {len(texts)} documents to vector store...")
                self.vector_store.add_documents(texts)
                
                logger.info(f"Successfully ingested {len(texts)} chunks from {file.filename}")
                return {
                    "success": True, 
                    "chunks": len(texts),
                    "filename": file.filename
                }

            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

        except Exception as e:
            logger.error(f"Error processing file {file.filename}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

    def _extract_xlsx_text(self, file_path: str) -> str:
        """
        Extract text from an Excel file using openpyxl.
        Converts each row into a descriptive string.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                all_text.append(f"Sheet: {sheet_name}")
                
                # Get column names from the first row
                headers = []
                try:
                    first_row = next(sheet.iter_rows(min_row=1, max_row=1))
                    headers = [str(cell.value) if cell.value is not None else "" for cell in first_row]
                except StopIteration:
                    continue
                
                # Iterate rows starting from second
                for row in sheet.iter_rows(min_row=2):
                    row_parts = []
                    for col_idx, cell in enumerate(row):
                        val = cell.value
                        if val is not None:
                            header = headers[col_idx] if col_idx < len(headers) and headers[col_idx] else f"Col{col_idx+1}"
                            str_val = str(val).strip()
                            if str_val:
                                row_parts.append(f"{header}: {str_val}")
                    
                    if row_parts:
                        all_text.append(", ".join(row_parts))
            
            return "\n\n".join(all_text)
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise ValueError(f"Could not read Excel file: {str(e)}")

ingest_service = IngestService()
